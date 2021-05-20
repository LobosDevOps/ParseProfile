import glob
import os

from openpyxl import Workbook

from utils import utils

def getXmlDataByList(path, note, subNotes, withns, sfdc_metadata):
    resultData = []
    for fl in glob.glob(path + "/**/*", recursive=True):
        if withns:
            xmldata = utils.parseXMLWithns(fl, subNotes, note, sfdc_metadata)
        else:
            xmldata = utils.parseXMLWithoutNs(fl, subNotes, note)
        # print(xmldata)
        resultData.extend(xmldata)
    # print('getXmlDataByList resultData : ', resultData)
    return resultData

# Excel Common
def addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes):
    sheet = wb.create_sheet(title=note)

    subNotes.insert(0, 'filename')
    # sheet header
    for i, objName in enumerate(subNotes, start=1):
        sheet.cell(1, i, value=objName)

    # sheet data
    resultData = getXmlDataByList(inputdir, note, subNotes, withns, sfdc_metadata)
    print(resultData)
    for rowNum, objName in enumerate(resultData, start=2):
        for columnNum, noteName in enumerate(subNotes, start=1):
            if (noteName in objName):
                cellValue = objName[noteName]
            else:
                cellValue = ''

            sheet.cell(row=rowNum, column=columnNum).value = cellValue

    return sheet

def getConfig():
    configObj = utils.getConfigInfo('permissionset.yaml')
    configMap = {}
    configMap['inputdir'] = configObj['inputdir']
    configMap['outputFileName'] = configObj['outputFileName']
    configMap['sfdc_metadata'] = configObj['sfdc_metadata']
    configMap['withns'] = configObj['withns']

    # label
    configMap['label_isTarget'] = configObj['field']['label']['isTarget']
    configMap['label_note'] = configObj['field']['label']['note']
    configMap['label_subNotes'] = configObj['field']['label']['subNotes']
    # license
    configMap['license_isTarget'] = configObj['field']['license']['isTarget']
    configMap['license_note'] = configObj['field']['license']['note']
    configMap['license_subNotes'] = configObj['field']['license']['subNotes']
    # hasActivationRequired
    configMap['actReq_isTarget'] = configObj['field']['hasActivationRequired']['isTarget']
    configMap['actReq_note'] = configObj['field']['hasActivationRequired']['note']
    configMap['actReq_subNotes'] = configObj['field']['hasActivationRequired']['subNotes']
    # applicationVisibilities
    configMap['app_isTarget'] = configObj['field']['applicationVisibilities']['isTarget']
    configMap['app_note'] = configObj['field']['applicationVisibilities']['note']
    configMap['app_subNotes'] = configObj['field']['applicationVisibilities']['subNotes']

    # classAccesses
    configMap['class_isTarget'] = configObj['field']['classAccesses']['isTarget']
    configMap['class_note'] = configObj['field']['classAccesses']['note']
    configMap['class_subNotes'] = configObj['field']['classAccesses']['subNotes']
    # customMetadataTypeAccesses
    configMap['metaType_isTarget'] = configObj['field']['customMetadataTypeAccesses']['isTarget']
    configMap['metaType_note'] = configObj['field']['customMetadataTypeAccesses']['note']
    configMap['metaType_subNotes'] = configObj['field']['customMetadataTypeAccesses']['subNotes']
    # customPermissions
    configMap['ctsPms_isTarget'] = configObj['field']['customPermissions']['isTarget']
    configMap['ctsPms_note'] = configObj['field']['customPermissions']['note']
    configMap['ctsPms_subNotes'] = configObj['field']['customPermissions']['subNotes']
    # customSettingAccesses
    configMap['ctsSet_isTarget'] = configObj['field']['customSettingAccesses']['isTarget']
    configMap['ctsSet_note'] = configObj['field']['customSettingAccesses']['note']
    configMap['ctsSet_subNotes'] = configObj['field']['customSettingAccesses']['subNotes']
    # externalDataSourceAccesses
    configMap['dataSrc_isTarget'] = configObj['field']['externalDataSourceAccesses']['isTarget']
    configMap['dataSrc_note'] = configObj['field']['externalDataSourceAccesses']['note']
    configMap['dataSrc_subNotes'] = configObj['field']['externalDataSourceAccesses']['subNotes']


    # fieldPermissions
    configMap['fldPms_isTarget'] = configObj['field']['fieldPermissions']['isTarget']
    configMap['fldPms_note'] = configObj['field']['fieldPermissions']['note']
    configMap['fldPms_subNotes'] = configObj['field']['fieldPermissions']['subNotes']

    # flowAccesses
    configMap['flow_isTarget'] = configObj['field']['flowAccesses']['isTarget']
    configMap['flow_note'] = configObj['field']['flowAccesses']['note']
    configMap['flow_subNotes'] = configObj['field']['flowAccesses']['subNotes']
    # objectPermissions
    configMap['obj_isTarget'] = configObj['field']['objectPermissions']['isTarget']
    configMap['obj_note'] = configObj['field']['objectPermissions']['note']
    configMap['obj_subNotes'] = configObj['field']['objectPermissions']['subNotes']
    # pageAccesses
    configMap['page_isTarget'] = configObj['field']['pageAccesses']['isTarget']
    configMap['page_note'] = configObj['field']['pageAccesses']['note']
    configMap['page_subNotes'] = configObj['field']['pageAccesses']['subNotes']
    # recordTypeVisibilities
    configMap['rt_isTarget'] = configObj['field']['recordTypeVisibilities']['isTarget']
    configMap['rt_note'] = configObj['field']['recordTypeVisibilities']['note']
    configMap['rt_subNotes'] = configObj['field']['recordTypeVisibilities']['subNotes']
    # tabSettings
    configMap['tab_isTarget'] = configObj['field']['tabSettings']['isTarget']
    configMap['tab_note'] = configObj['field']['tabSettings']['note']
    configMap['tab_subNotes'] = configObj['field']['tabSettings']['subNotes']
    # userPermissions
    configMap['usrPms_isTarget'] = configObj['field']['userPermissions']['isTarget']
    configMap['usrPms_note'] = configObj['field']['userPermissions']['note']
    configMap['usrPms_subNotes'] = configObj['field']['userPermissions']['subNotes']

    # print('configMap', configMap)
    return configMap

def outputXmlDataToExcel():
    configMap = getConfig()
    inputdir = configMap['inputdir']
    sfdc_metadata = configMap['sfdc_metadata']
    withns = configMap['withns']

    wb = Workbook()

    # label
    addLabelSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # license
    addLicenseSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # hasActivationRequired
    addActReqSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # applicationVisibilities
    addAppSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # classAccesses
    addClassSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # customMetadataTypeAccesses
    addMetaTypeSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # customPermissions
    addCtsPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # customSettingAccesses
    addCtsSetSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # externalDataSourceAccesses
    addDataSrcSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # fieldPermissions
    addFldPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # flowAccesses
    addFlowSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # objectPermissions
    addObjSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # pageAccesses
    addPageSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # recordTypeVisibilities
    addRtSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # tabSettings
    addTabSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # userPermissions
    addUsrPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns)

    if(len(wb.sheetnames) > 1):
        wb.remove(wb.worksheets[0])

    print(wb.sheetnames)
    wb.save(configMap['outputFileName'] + '.xlsx')

# label
def addLabelSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['label_isTarget']
    note = configMap['label_note']
    subNotes = configMap['label_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# license
def addLicenseSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['license_isTarget']
    note = configMap['license_note']
    subNotes = configMap['license_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# hasActivationRequired
def addActReqSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['actReq_isTarget']
    note = configMap['actReq_note']
    subNotes = configMap['actReq_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# applicationVisibilities
def addAppSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['app_isTarget']
    note = configMap['app_note']
    subNotes = configMap['app_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# classAccesses
def addClassSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['class_isTarget']
    note = configMap['class_note']
    subNotes = configMap['class_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# customMetadataTypeAccesses
def addMetaTypeSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['metaType_isTarget']
    note = configMap['metaType_note']
    subNotes = configMap['metaType_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# customPermissions
def addCtsPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['ctsPms_isTarget']
    note = configMap['ctsPms_note']
    subNotes = configMap['ctsPms_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# customSettingAccesses
def addCtsSetSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['ctsSet_isTarget']
    note = configMap['ctsSet_note']
    subNotes = configMap['ctsSet_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# externalDataSourceAccesses
def addDataSrcSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['dataSrc_isTarget']
    note = configMap['dataSrc_note']
    subNotes = configMap['dataSrc_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# fieldPermissions
def addFldPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['fldPms_isTarget']
    note = configMap['fldPms_note']
    subNotes = configMap['fldPms_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# flowAccesses
def addFlowSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['flow_isTarget']
    note = configMap['flow_note']
    subNotes = configMap['flow_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# objectPermissions
def addObjSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['obj_isTarget']
    note = configMap['obj_note']
    subNotes = configMap['obj_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# pageAccesses
def addPageSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['page_isTarget']
    note = configMap['page_note']
    subNotes = configMap['page_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# recordTypeVisibilities
def addRtSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['rt_isTarget']
    note = configMap['rt_note']
    subNotes = configMap['rt_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# tabSettings
def addTabSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['tab_isTarget']
    note = configMap['tab_note']
    subNotes = configMap['tab_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# userPermissions
def addUsrPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['usrPms_isTarget']
    note = configMap['usrPms_note']
    subNotes = configMap['usrPms_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)