import glob
import os

from openpyxl import Workbook

from utils import utils

def getXmlDataByList(path, note, subNotes, withns, sfdc_metadata):
    resultData = []
    for fl in glob.glob(path + "/**/*", recursive=True):
        if withns:
            xmldata = utils.parseXMLWithns(fl, subNotes, note, sfdc_metadata)
        else:
            xmldata = utils.parseXMLWithoutNs(fl, subNotes, note)
        # print(xmldata)
        resultData.extend(xmldata)
    # print('getXmlDataByList resultData : ', resultData)
    return resultData

# Excel Common
def addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes):
    sheet = wb.create_sheet(title=note)

    subNotes.insert(0, 'filename')
    # sheet header
    for i, objName in enumerate(subNotes, start=1):
        sheet.cell(1, i, value=objName)

    # sheet data
    resultData = getXmlDataByList(inputdir, note, subNotes, withns, sfdc_metadata)
    print(resultData)
    for rowNum, objName in enumerate(resultData, start=2):
        for columnNum, noteName in enumerate(subNotes, start=1):
            if (noteName in objName):
                cellValue = objName[noteName]
            else:
                cellValue = ''

            sheet.cell(row=rowNum, column=columnNum).value = cellValue

    return sheet

def getConfig():
    configObj = utils.getConfigInfo('profile.yaml')
    configMap = {}
    configMap['inputdir'] = configObj['inputdir']
    configMap['outputFileName'] = configObj['outputFileName']
    configMap['sfdc_metadata'] = configObj['sfdc_metadata']
    configMap['withns'] = configObj['withns']

    # # fullName
    # configMap['fullname_isTarget'] = configObj['field']['fullName']['isTarget']
    # configMap['fullname_note'] = configObj['field']['fullName']['note']
    # configMap['fullname_subNotes'] = configObj['field']['fullName']['subNotes']
    # userLicense
    configMap['license_isTarget'] = configObj['field']['userLicense']['isTarget']
    configMap['license_note'] = configObj['field']['userLicense']['note']
    configMap['license_subNotes'] = configObj['field']['userLicense']['subNotes']
    # custom
    configMap['custom_isTarget'] = configObj['field']['custom']['isTarget']
    configMap['custom_note'] = configObj['field']['custom']['note']
    configMap['custom_subNotes'] = configObj['field']['custom']['subNotes']
    # applicationVisibilities
    configMap['app_isTarget'] = configObj['field']['applicationVisibilities']['isTarget']
    configMap['app_note'] = configObj['field']['applicationVisibilities']['note']
    configMap['app_subNotes'] = configObj['field']['applicationVisibilities']['subNotes']
    # categoryGroupVisibilities
    configMap['ctg_isTarget'] = configObj['field']['categoryGroupVisibilities']['isTarget']
    configMap['ctg_note'] = configObj['field']['categoryGroupVisibilities']['note']
    configMap['ctg_subNotes'] = configObj['field']['categoryGroupVisibilities']['subNotes']
    # classAccesses
    configMap['class_isTarget'] = configObj['field']['classAccesses']['isTarget']
    configMap['class_note'] = configObj['field']['classAccesses']['note']
    configMap['class_subNotes'] = configObj['field']['classAccesses']['subNotes']
    # customMetadataTypeAccesses
    configMap['metaType_isTarget'] = configObj['field']['customMetadataTypeAccesses']['isTarget']
    configMap['metaType_note'] = configObj['field']['customMetadataTypeAccesses']['note']
    configMap['metaType_subNotes'] = configObj['field']['customMetadataTypeAccesses']['subNotes']
    # customPermissions
    configMap['ctsPms_isTarget'] = configObj['field']['customPermissions']['isTarget']
    configMap['ctsPms_note'] = configObj['field']['customPermissions']['note']
    configMap['ctsPms_subNotes'] = configObj['field']['customPermissions']['subNotes']
    # customSettingAccesses
    configMap['ctsSet_isTarget'] = configObj['field']['customSettingAccesses']['isTarget']
    configMap['ctsSet_note'] = configObj['field']['customSettingAccesses']['note']
    configMap['ctsSet_subNotes'] = configObj['field']['customSettingAccesses']['subNotes']
    # externalDataSourceAccesses
    configMap['dataSrc_isTarget'] = configObj['field']['externalDataSourceAccesses']['isTarget']
    configMap['dataSrc_note'] = configObj['field']['externalDataSourceAccesses']['note']
    configMap['dataSrc_subNotes'] = configObj['field']['externalDataSourceAccesses']['subNotes']
    # fieldPermissions
    configMap['fldPms_isTarget'] = configObj['field']['fieldPermissions']['isTarget']
    configMap['fldPms_note'] = configObj['field']['fieldPermissions']['note']
    configMap['fldPms_subNotes'] = configObj['field']['fieldPermissions']['subNotes']
    # flowAccesses
    configMap['flow_isTarget'] = configObj['field']['flowAccesses']['isTarget']
    configMap['flow_note'] = configObj['field']['flowAccesses']['note']
    configMap['flow_subNotes'] = configObj['field']['flowAccesses']['subNotes']

    # layoutAssignments
    configMap['layout_isTarget'] = configObj['field']['layoutAssignments']['isTarget']
    configMap['layout_note'] = configObj['field']['layoutAssignments']['note']
    configMap['layout_subNotes'] = configObj['field']['layoutAssignments']['subNotes']
    # loginFlows
    configMap['loginFlow_isTarget'] = configObj['field']['loginFlows']['isTarget']
    configMap['loginFlow_note'] = configObj['field']['loginFlows']['note']
    configMap['loginFlow_subNotes'] = configObj['field']['loginFlows']['subNotes']
    # loginHours
    configMap['loginHour_isTarget'] = configObj['field']['loginHours']['isTarget']
    configMap['loginHour_note'] = configObj['field']['loginHours']['note']
    configMap['loginHour_subNotes'] = configObj['field']['loginHours']['subNotes']
    # loginIpRanges
    configMap['loginIp_isTarget'] = configObj['field']['loginIpRanges']['isTarget']
    configMap['loginIp_note'] = configObj['field']['loginIpRanges']['note']
    configMap['loginIp_subNotes'] = configObj['field']['loginIpRanges']['subNotes']
    # objectPermissions
    configMap['obj_isTarget'] = configObj['field']['objectPermissions']['isTarget']
    configMap['obj_note'] = configObj['field']['objectPermissions']['note']
    configMap['obj_subNotes'] = configObj['field']['objectPermissions']['subNotes']
    # pageAccesses
    configMap['page_isTarget'] = configObj['field']['pageAccesses']['isTarget']
    configMap['page_note'] = configObj['field']['pageAccesses']['note']
    configMap['page_subNotes'] = configObj['field']['pageAccesses']['subNotes']
    # profileActionOverrides
    configMap['actionOverride_isTarget'] = configObj['field']['profileActionOverrides']['isTarget']
    configMap['actionOverride_note'] = configObj['field']['profileActionOverrides']['note']
    configMap['actionOverride_subNotes'] = configObj['field']['profileActionOverrides']['subNotes']
    # recordTypeVisibilities
    configMap['rt_isTarget'] = configObj['field']['recordTypeVisibilities']['isTarget']
    configMap['rt_note'] = configObj['field']['recordTypeVisibilities']['note']
    configMap['rt_subNotes'] = configObj['field']['recordTypeVisibilities']['subNotes']
    # tabVisibilities
    configMap['tab_isTarget'] = configObj['field']['tabVisibilities']['isTarget']
    configMap['tab_note'] = configObj['field']['tabVisibilities']['note']
    configMap['tab_subNotes'] = configObj['field']['tabVisibilities']['subNotes']
    # userPermissions
    configMap['usrPms_isTarget'] = configObj['field']['userPermissions']['isTarget']
    configMap['usrPms_note'] = configObj['field']['userPermissions']['note']
    configMap['usrPms_subNotes'] = configObj['field']['userPermissions']['subNotes']

    # print('configMap', configMap)
    return configMap

def outputXmlDataToExcel():
    configMap = getConfig()
    inputdir = configMap['inputdir']
    sfdc_metadata = configMap['sfdc_metadata']
    withns = configMap['withns']

    wb = Workbook()
    # # fullName
    # addFullNameSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # userLicense
    addLicenseSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # custom
    addCustomSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # applicationVisibilities
    addAppSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # categoryGroupVisibilities
    addCtgSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # classAccesses
    addClassSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # customMetadataTypeAccesses
    addMetaTypeSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # customPermissions
    addCtsPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # customSettingAccesses
    addCtsSetSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # externalDataSourceAccesses
    addDataSrcSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # fieldPermissions
    addFldPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # flowAccesses
    addFlowSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # layoutAssignments
    addLayoutSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # loginFlows
    addLoginFlowSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # loginHours
    addLoginHourSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # loginIpRanges
    addLoginIpSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # objectPermissions
    addObjSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # pageAccesses
    addPageSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # profileActionOverrides
    addActionOverrideSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # recordTypeVisibilities
    addRtSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # tabSettings
    addTabSheet(wb, configMap, inputdir, sfdc_metadata, withns)
    # userPermissions
    addUsrPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns)

    if(len(wb.sheetnames) > 1):
        wb.remove(wb.worksheets[0])

    print(wb.sheetnames)
    wb.save(configMap['outputFileName'] + '.xlsx')

# # fullName
# def addFullNameSheet(wb, configMap, inputdir, sfdc_metadata, withns):
#     isTarget = configMap['fullname_isTarget']
#     note = configMap['fullname_note']
#     subNotes = configMap['fullname_subNotes']

#     if(isTarget == False):
#         return

#     addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# userLicense
def addLicenseSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['license_isTarget']
    note = configMap['license_note']
    subNotes = configMap['license_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# custom
def addCustomSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['custom_isTarget']
    note = configMap['custom_note']
    subNotes = configMap['custom_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# applicationVisibilities
def addAppSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['app_isTarget']
    note = configMap['app_note']
    subNotes = configMap['app_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# categoryGroupVisibilities
def addCtgSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['ctg_isTarget']
    note = configMap['ctg_note']
    subNotes = configMap['ctg_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# classAccesses
def addClassSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['class_isTarget']
    note = configMap['class_note']
    subNotes = configMap['class_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# customMetadataTypeAccesses
def addMetaTypeSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['metaType_isTarget']
    note = configMap['metaType_note']
    subNotes = configMap['metaType_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# customPermissions
def addCtsPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['ctsPms_isTarget']
    note = configMap['ctsPms_note']
    subNotes = configMap['ctsPms_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# customSettingAccesses
def addCtsSetSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['ctsSet_isTarget']
    note = configMap['ctsSet_note']
    subNotes = configMap['ctsSet_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# externalDataSourceAccesses
def addDataSrcSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['dataSrc_isTarget']
    note = configMap['dataSrc_note']
    subNotes = configMap['dataSrc_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# fieldPermissions
def addFldPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['fldPms_isTarget']
    note = configMap['fldPms_note']
    subNotes = configMap['fldPms_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# flowAccesses
def addFlowSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['flow_isTarget']
    note = configMap['flow_note']
    subNotes = configMap['flow_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# layoutAssignments
def addLayoutSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['layout_isTarget']
    note = configMap['layout_note']
    subNotes = configMap['layout_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# loginFlows
def addLoginFlowSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['loginFlow_isTarget']
    note = configMap['loginFlow_note']
    subNotes = configMap['loginFlow_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# loginHours
def addLoginHourSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['loginHour_isTarget']
    note = configMap['loginHour_note']
    subNotes = configMap['loginHour_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# loginIpRanges
def addLoginIpSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['loginIp_isTarget']
    note = configMap['loginIp_note']
    subNotes = configMap['loginIp_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# objectPermissions
def addObjSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['obj_isTarget']
    note = configMap['obj_note']
    subNotes = configMap['obj_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# pageAccesses
def addPageSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['page_isTarget']
    note = configMap['page_note']
    subNotes = configMap['page_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# profileActionOverrides
def addActionOverrideSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['actionOverride_isTarget']
    note = configMap['actionOverride_note']
    subNotes = configMap['actionOverride_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# recordTypeVisibilities
def addRtSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['rt_isTarget']
    note = configMap['rt_note']
    subNotes = configMap['rt_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# tabSettings
def addTabSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['tab_isTarget']
    note = configMap['tab_note']
    subNotes = configMap['tab_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)

# userPermissions
def addUsrPmsSheet(wb, configMap, inputdir, sfdc_metadata, withns):
    isTarget = configMap['usrPms_isTarget']
    note = configMap['usrPms_note']
    subNotes = configMap['usrPms_subNotes']

    if(isTarget == False):
        return

    addExcelSheet(wb, inputdir, sfdc_metadata, withns, note, subNotes)