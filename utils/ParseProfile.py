import glob
import os

from openpyxl import Workbook

from utils import utils

def getXmlDataByList(path, targetNote, targetSubNotes, withns, sfdc_metadata):
    resultData = []
    for fl in glob.glob(path + "/**/*", recursive=True):
        if withns:
            xmldata = utils.parseXMLWithns(fl, targetSubNotes, targetNote, sfdc_metadata)
        else:
            xmldata = utils.parseXMLWithoutNs(fl, targetSubNotes, targetNote)
        # print(xmldata)
        resultData.extend(xmldata)
    # print('getXmlDataByList resultData : ', resultData)
    return resultData

def getXmlDataToCsvByMatrix(path, targetNote, targetSubNotes, withns, sfdc_metadata, targetSubNotekey):
    xmldata = getXmlDataByList(path, targetNote, targetSubNotes, withns, sfdc_metadata)
    fileDataMap = {}

    for xmlDataMap in xmldata:
        print('xmlDataMap:', xmlDataMap)
        index = 0
        cellKey = ''
        cellValue = ''
        for key, value in xmlDataMap.items():
            # print(index, key, value)
            if (index == 0):
                if (not value in fileDataMap):
                    notes = {}
                    notes[key] = value
                    fileDataMap[value] = notes
                    csvHead = key
            else:
                if (targetSubNotekey == key):
                    csvHead += ',' + value
                    cellKey = value
                    # print('notes :', notes)
                else:
                    cellValue = value

            # print(fileDataMap.values())
            index += 1
        notes[cellKey] = cellValue

    # print(csvHead)
    # print(fileDataMap.values())
    resultMap = {'head': csvHead, 'datas': list(fileDataMap.values())}
    print('resultMap : ', resultMap)
    return resultMap

def getConfig(configObj):
    configMap = {}
    configMap['inputdir'] = configObj['inputdir']
    configMap['outputFileName'] = configObj['outputFileName']
    configMap['sfdc_metadata'] = configObj['sfdc_metadata']
    configMap['withns'] = configObj['withns']

    # csv
    configMap['csv_targetNote'] = configObj['csv']['targetNote']
    configMap['csv_targetSubNotes'] = configObj['csv']['targetSubNotes']
    configMap['csv_targetKey'] = configObj['csv']['targetKey']

    # For Excel(Matrix)
    # userPermissions
    configMap['exl_usrPms_targetNote'] = configObj['exl']['userPermissions']['targetNote']
    configMap['exl_usrPms_load'] = configObj['exl']['userPermissions']['load']
    configMap['exl_usrPms_targetSubNotes'] = configObj['exl']['userPermissions']['targetSubNotes']
    configMap['exl_usrPms_targetKey'] = configObj['exl']['userPermissions']['targetKey']
    configMap['exl_usrPms_targetObjs'] = configObj['exl']['userPermissions']['targetObjs']

    # tabVisibilities
    configMap['exl_tab_targetNote'] = configObj['exl']['tabVisibilities']['targetNote']
    configMap['exl_tab_load'] = configObj['exl']['tabVisibilities']['load']
    configMap['exl_tab_targetSubNotes'] = configObj['exl']['tabVisibilities']['targetSubNotes']
    configMap['exl_tab_targetKey'] = configObj['exl']['tabVisibilities']['targetKey']
    configMap['exl_tab_targetObjs'] = configObj['exl']['tabVisibilities']['targetObjs']

    print('configMap', configMap)
    return configMap

def outputXmlDataToCsvByList(configObj, isOutputFile):
    configMap = getConfig(configObj)

    path = configMap['inputdir']
    sfdc_metadata = configMap['sfdc_metadata']
    withns = configMap['withns']

    # csv
    targetNote = configMap['csv_targetNote']
    targetSubNotes = configMap['csv_targetSubNotes']

    outputFileName = configMap['outputFileName'] + '_' + path.split('/')[- 1] + '_' + targetNote + 'List.csv'
    xmldata = getXmlDataByList(path, targetNote, targetSubNotes, withns, sfdc_metadata)

    if isOutputFile:
        if os.path.exists(outputFileName):
            os.remove(outputFileName)
        utils.savetoCSV(targetSubNotes, xmldata, outputFileName)

def outputXmlDataToCsvByMatrix(configObj, isOutputFile):
    configMap = getConfig(configObj)
    path = configMap['inputdir']
    sfdc_metadata = configMap['sfdc_metadata']
    withns = configMap['withns']

    # csv
    targetNote = configMap['csv_targetNote']
    targetSubNotes = configMap['csv_targetSubNotes']
    targetKey = configMap['csv_targetKey']

    outputFileName = configMap['outputFileName'] + '_' + path.split('/')[- 1] + '_' + targetNote + 'Matrix.csv'

    dataMap = getXmlDataToCsvByMatrix(path, targetNote, targetSubNotes, withns, sfdc_metadata, targetKey)
    if isOutputFile:
        if os.path.exists(outputFileName):
            os.remove(outputFileName)
        utils.savetoCSV(dataMap['head'].split(','), dataMap['datas'], outputFileName)

# Excel Common
def addExcelSheet(wb, path, sfdc_metadata, withns, targetNote, targetSubNotes, targetKey, targetObjs):
    targetObjs.insert(0, 'filename')
    # print('targetObjs:', targetObjs)

    dataMap = getXmlDataToCsvByMatrix(path, targetNote, targetSubNotes, withns, sfdc_metadata, targetKey)
    # print(dataMap)
    datas = dataMap['datas']
    # print('list(datas):', datas)

    sheet = wb.create_sheet(title=targetNote)

    for i, objName in enumerate(targetObjs, start=0):
        sheet.cell(i+1, 1, value=objName)

    for rowNum, datas in enumerate(datas, start=2):
        print('datas', datas)
        for columnNum, objName in enumerate(targetObjs, start=1):
            if (objName in datas):
                print(rowNum, columnNum, objName, datas[objName])
                cellValue = datas[objName]
            else:
                print(rowNum, columnNum, 'none')
                cellValue = ''
            sheet.cell(row=columnNum, column=rowNum).value = cellValue

    return sheet

# userPermissions
def addUserPermissionsSheet(wb, configMap, path, sfdc_metadata, withns):
    # userPermissions
    targetNote = configMap['exl_usrPms_targetNote']
    load = configMap['exl_usrPms_load']
    targetSubNotes = configMap['exl_usrPms_targetSubNotes']
    targetKey = configMap['exl_usrPms_targetKey']
    targetObjs = configMap['exl_usrPms_targetObjs']

    if(load == False):
        return

    addExcelSheet(wb, path, sfdc_metadata, withns, targetNote, targetSubNotes, targetKey, targetObjs)

# tabVisibilities
def addTabVisibilitiesSheet(wb, configMap, path, sfdc_metadata, withns):
    # tabVisibilities
    targetNote = configMap['exl_tab_targetNote']
    load = configMap['exl_tab_load']
    targetSubNotes = configMap['exl_tab_targetSubNotes']
    targetKey = configMap['exl_tab_targetKey']
    targetObjs = configMap['exl_tab_targetObjs']
    if(load == False):
        return

    addExcelSheet(wb, path, sfdc_metadata, withns, targetNote, targetSubNotes, targetKey, targetObjs)


def outputXmlDataToExcelByMatrix(configObj, isOutputFile):
    configMap = getConfig(configObj)
    path = configMap['inputdir']
    sfdc_metadata = configMap['sfdc_metadata']
    withns = configMap['withns']

    wb = Workbook()

    # userPermissions
    addUserPermissionsSheet(wb, configMap, path, sfdc_metadata, withns)

    # tabVisibilities
    addTabVisibilitiesSheet(wb, configMap, path, sfdc_metadata, withns)

    print(wb.sheetnames)
    if(len(wb.sheetnames) > 1):
        wb.remove(wb.worksheets[0])
    wb.save('testリスト追加3.xlsx')

    # for rowNum, datas in enumerate(datas, start=2):
    #     for columnNum, (key, value) in enumerate(datas.items(), start=1):
    #         print(rowNum, columnNum, key, value)
    #         # sheet.cell(row=rowNum, column=columnNum).value = value

    # wb.save('testリスト追加.xlsx')



    # for rowNum, datas in enumerate(datas, start=2):
    #     for columnNum, (key, value) in enumerate(datas.items(), start=1):
    #         print(rowNum, columnNum, key, value)
    #         sheet.cell(row=rowNum, column=columnNum).value = value

    # dataList = list(datas)
    # # print(dataList)
    # for rowNum, datas in enumerate(dataList, start=2):
    #     for columnNum, (key, value) in enumerate(datas.items(), start=1):
    #         print(rowNum, columnNum, key, value)
    #         sheet.cell(row=rowNum, column=columnNum).value = value


    # for row, data in enumerate(dataList, start=2):
    #     sheet[f"A{row}"] = data['filename']
    #     sheet[f"B{row}"] = data['Account']
    #     sheet[f"C{row}"] = data['ActivityStatus__c']
    # wb.save('testリスト追加.xlsx')
